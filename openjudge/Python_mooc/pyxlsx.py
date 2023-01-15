import openpyxl as xls
import datetime
book = xls.Workbook()
sheet = book.active
sheet.title = "sample1"
dataRows = ((10, 20, 30, 40.5), 
    (100, 200, '=sum(A1:B2)'), 
    [],
    ['1000', datetime.datetime.now(), 'ok'])
for row in dataRows:
    sheet.append(row)
sheet.column_dimensions['B'].width = len(str(sheet['B4'].value))
sheet['E1'].value = '=sum(A1:D1)'
sheet['E2'].value = 12.5
sheet['E2'].number_format = '0.00%'
sheet['F1'].value = 3500
sheet['F2'].value = "35.00"
sheet['F3'].value = datetime.datetime.today().date()
sheet.column_dimensions['F'].width = len(str(sheet['F3'].value))
sheet.row_dimensions[2].height = 100

sheet2 = book.create_sheet("Sample2")
sheet2['A1'] = 50
sheet2 = book.create_sheet("Sample0", 0)
sheet3 = book.copy_worksheet(sheet)
# book.remove()
book.save('./tmp.xlsx')