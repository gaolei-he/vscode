from datetime import date, datetime
import openpyxl as xls
book = xls.load_workbook('')
sheet = book.worksheets[0]
# sheet = book.active
# sheet = book['price']
# for sheet in book.worksheets:
    # print(sheet.title)
print(sheet.title)
print(sheet.min_row, sheet.max_row)
print(sheet.min_column, sheet.max_row)
for row in sheet.rows:
    for cell in row:
        print(cell.value)
for cell in sheet['G']:
    print(cell.value)
for cell in sheet[3]:
    print(cell.value, type(cell.value), cell.coordinate,
    cell.col_idx, cell.number_format)


# type(cell.value) : int, float, str, datetime.datetime
# cell.coordinate : 'A2', 'E3'
# cell.col_idx : 单元格列号
# cell.number_format : 数的显示格式， 'General', '0.00%', '0.00E+00'