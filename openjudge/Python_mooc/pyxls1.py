import shelve
import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Alignment, Side, Border
book = openpyxl.Workbook()
sheet = book.active
for i in range(4):
    sheet.append([i*5+j for j in range(5)])
side = Side(style='thin')
border = Border(left=side, right=side, top=side, bottom=side)
for row in sheet.rows:
    for cell in row:
        cell.border = border
sheet['A1'].fill = PatternFill(patternType='solid', start_color=colors.BLACK)
a1 = sheet['A1']
italicRedFont = Font(size=18, name='Times New Roman', bold=True, color=colors.BLUE)
a1.font = italicRedFont
sheet['A2'].font = sheet['A1'].font.copy(italic=True)

sheet.merge_cells('C2:D3')
sheet['C2'].alignment = Alignment(horizontal='left', vertical='center')
book.save('./tmp.xlsx')